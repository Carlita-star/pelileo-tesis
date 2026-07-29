from typing import Any, List, Optional

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _format_now() -> str:
    return timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')


def _column_letter_range(col_count: int) -> str:
    if col_count <= 0:
        return 'A'
    return get_column_letter(col_count)


def _auto_column_widths(headers: List[str], rows: List[List[Any]]) -> List[float]:
    widths = []
    for col_idx, header in enumerate(headers):
        max_len = len(str(header))
        for row in rows:
            if col_idx < len(row):
                max_len = max(max_len, len(str(row[col_idx] or '')))
        widths.append(min(max(max_len + 2, 10), 48))
    return widths


def generar_excel_formateado(
    titulo: str,
    headers: List[str],
    rows: List[List[Any]],
    filepath: str,
    subtitulo: Optional[str] = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    col_count = max(len(headers), 1)
    last_col = _column_letter_range(col_count)

    thin = Side(style='thin', color='D0D8E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1D74F2')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='122540')
    meta_font = Font(size=10, color='6F7A95')
    alt_fill = PatternFill('solid', fgColor='F7FBFF')

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    meta_text = subtitulo or f'Generado: {_format_now()}'
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = meta_text
    ws['A2'].font = meta_font
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = f'Total de registros: {len(rows)}'
    ws['A3'].font = meta_font

    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 24

    if not rows:
        cell = ws.cell(
            row=header_row + 1,
            column=1,
            value='Sin registros para los filtros seleccionados.',
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 1,
            end_column=col_count,
        )
    else:
        for row_idx, row in enumerate(rows, start=header_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    for index, width in enumerate(_auto_column_widths(headers, rows), start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    wb.save(filepath)


def _pdf_cell(value: Any, styles) -> Paragraph:
    text = '' if value is None else str(value)
    safe = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return Paragraph(safe or '—', styles['Cell'])


def generar_pdf_formateado(
    titulo: str,
    headers: List[str],
    rows: List[List[Any]],
    filepath: str,
    subtitulo: Optional[str] = None,
) -> None:
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Cell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        name='Meta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6F7A95'),
    ))

    elements = [
        Paragraph(titulo, styles['Title']),
        Paragraph(subtitulo or f'Generado: {_format_now()}', styles['Meta']),
        Paragraph(f'Total de registros: {len(rows)}', styles['Meta']),
        Spacer(1, 10),
    ]

    if not rows:
        table_data = [
            headers,
            ['Sin registros para los filtros seleccionados.'] + [''] * (len(headers) - 1),
        ]
    else:
        table_data = [headers] + [
            [_pdf_cell(cell, styles) for cell in row]
            for row in rows
        ]

    available_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    col_width = available_width / max(len(headers), 1)
    col_widths = [col_width] * len(headers)

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D74F2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D8E8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FBFF')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)


def build_filtros_subtitulo(filtros: Optional[dict]) -> str:
    filtros = filtros or {}
    partes = [f'Generado: {_format_now()}']

    estado = (filtros.get('estado') or '').strip()
    if estado and estado.lower() != 'todos':
        partes.append(f'Estado: {estado.capitalize()}')

    categoria_id = filtros.get('categoria_id')
    categoria_nombre = filtros.get('categoria_nombre')
    if categoria_nombre:
        partes.append(f'Categoría: {categoria_nombre}')
    elif categoria_id:
        partes.append(f'Categoría ID: {categoria_id}')

    desde = filtros.get('desde')
    hasta = filtros.get('hasta')
    if desde or hasta:
        partes.append(f'Periodo: {desde or "…"} — {hasta or "…"}')

    return ' | '.join(partes)
