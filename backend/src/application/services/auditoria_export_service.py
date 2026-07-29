import csv
from io import BytesIO, StringIO
from typing import Iterable, List

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.domain.auditorias.entities import AuditoriaEntity

HEADERS = ['Fecha y hora', 'Usuario', 'Acción', 'Módulo', 'ID registro', 'IP']

MODULO_LABELS = {
    'atractivos': 'Atractivos',
    'rutas': 'Rutas',
    'emprendimientos': 'Emprendimientos',
    'usuarios': 'Usuarios',
    'eventos': 'Eventos',
    'configuracion': 'Configuración',
    'catalogos': 'Catálogos',
}


def _format_fecha(dt) -> str:
    if not dt:
        return ''
    local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local.strftime('%d/%m/%Y %H:%M:%S')


def _format_modulo(tabla: str) -> str:
    if not tabla:
        return 'Sin módulo'
    key = tabla.strip().lower()
    return MODULO_LABELS.get(key, tabla.replace('_', ' ').title())


def _build_rows(registros: Iterable[AuditoriaEntity]) -> List[List]:
    rows = []
    for item in registros:
        rows.append([
            _format_fecha(item.fecha),
            item.nombre_usuario or 'Sistema',
            item.accion or '',
            _format_modulo(item.tabla_afectada),
            item.entidad_id if item.entidad_id is not None else '',
            item.ip_address or '',
        ])
    return rows


def auditorias_to_csv(registros: List[AuditoriaEntity]) -> str:
    """CSV compatible con Excel en español (UTF-8 BOM + separador ;)."""
    buffer = StringIO()
    buffer.write('\ufeff')
    writer = csv.writer(buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(HEADERS)
    for row in _build_rows(registros):
        writer.writerow(row)
    return buffer.getvalue()


def auditorias_to_xlsx(registros: List[AuditoriaEntity]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Auditoría'

    thin = Side(style='thin', color='D0D8E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1D74F2')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='122540')
    meta_font = Font(size=10, color='6F7A95')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Reporte de auditoría'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Generado: {_format_fecha(timezone.now())}'
    ws['A2'].font = meta_font

    header_row = 4
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    data_rows = _build_rows(registros)
    if not data_rows:
        cell = ws.cell(row=header_row + 1, column=1, value='Sin registros para los filtros seleccionados.')
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 1,
            end_column=len(HEADERS),
        )
    else:
        alt_fill = PatternFill('solid', fgColor='F7FBFF')
        for row_idx, row in enumerate(data_rows, start=header_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    column_widths = [20, 28, 14, 18, 12, 16]
    for index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
